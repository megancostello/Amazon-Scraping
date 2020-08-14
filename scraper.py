import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import datetime
import tkinter as tk

incorrectVariations =[]

def scrapeAmazon(startRow, endRow):
    workbook = load_workbook(filename="Log.xlsx")
    sheet = workbook["Logged"]

    headers1 = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.105 Safari/537.36'}
    rowCounter = startRow

    #function for getting Amazon page data via requests.get() and convert into BeautifulSoup obj
    def getData(pageURL):
        response = requests.get(pageURL, headers=headers1)
        soupy = BeautifulSoup(response.content, 'html.parser')
        return soupy

    #cycles through rows in column 6 which has URLs
    for urlRow in range(startRow, endRow):
        url = workbook["Actual"].cell(row=urlRow, column=9).value
        soup = getData(url)
        title = ''
        hasTitle = False
        tryCount = 5
        while hasTitle is not True:
            try:
                title = soup.find(id="productTitle").get_text().rstrip()
                hasTitle = True
                print('soup is good')
            except AttributeError or TypeError:
                if tryCount == 0:
                    break
                soup = getData(url)
                tryCount-=1
                print('soup redone')

        if tryCount != 0:

            brand = soup.find(id="bylineInfo").get_text().rstrip()
            brand = brand.replace('Visit the ', '')
            brand = brand.replace(' Store', '')

            styleCount = 0
            baseStyleName = ''
            dropdownStyle = False
            if soup.find(id="native_dropdown_selected_style_name") is not None:
                baseStyleName = "native_style_name_"
                dropdownStyle=True
            else:
                baseStyleName = "style_name_"
            styleName = baseStyleName + str(styleCount)

            while soup.find(id=styleName):
                styleCount+=1
                styleName = baseStyleName + str(styleCount)

            hasStyles = False
            if styleCount > 0:
                hasStyles = True

            sizeCount = 0
            baseSizeName = "size_name_"
            sizeName = baseSizeName + str(sizeCount)

            while soup.find(id=sizeName):
                sizeCount+=1
                sizeName = baseSizeName + str(sizeCount)

            hasSizes = False
            if sizeCount > 0:
                hasSizes = True

            colorCount = 0
            baseColorName = "color_name_"
            colorName = baseColorName + str(colorCount)

            while soup.find(id=colorName):
                colorCount+=1
                colorName = baseColorName + str(colorCount)

            hasColors = False
            if colorCount > 0:
                hasColors = True

            selectedStyle = ''
            selections = []

            if dropdownStyle:
                selectedStyle = soup.find(id="dropdown_selected_style_name").get_text().rstrip()
                print(selectedStyle)
                sheet.cell(row=rowCounter, column=6).value = selectedStyle
                hasStyles = False

            allSelections = soup.findAll("span", attrs={"class": "selection"})

            for selection in allSelections:
                temp = selection.get_text().rstrip()
                temp2 = temp.replace("\n","")
                selections.append(temp2)

            if hasStyles and hasSizes and hasColors:
                sheet.cell(row=rowCounter, column=6).value = selections[0]
                sheet.cell(row=rowCounter, column=7).value = selections[1]
                sheet.cell(row=rowCounter, column=8).value = selections[2]
            elif hasStyles and hasSizes and not hasColors:
                sheet.cell(row=rowCounter, column=7).value = selections[0]
                sheet.cell(row=rowCounter, column=6).value = selections[1]
                sheet.cell(row=rowCounter, column=8).value = None
            elif not hasStyles and hasSizes and hasColors:
                sheet.cell(row=rowCounter, column=7).value = selections[0]
                sheet.cell(row=rowCounter, column=8).value = selections[1]
                if not dropdownStyle:
                    sheet.cell(row=rowCounter, column=6).value = None
            elif hasStyles and not hasSizes and hasColors:
                sheet.cell(row=rowCounter, column=6).value = selections[0]
                sheet.cell(row=rowCounter, column=8).value = selections[1]
                sheet.cell(row=rowCounter, column=7).value = None
            elif hasStyles and not hasSizes and not hasColors:
                sheet.cell(row=rowCounter, column=6).value = selections[0]
                sheet.cell(row=rowCounter, column=7).value = None
                sheet.cell(row=rowCounter, column=8).value = None
            elif not hasStyles and hasSizes and not hasColors:
                sheet.cell(row=rowCounter, column=7).value = selections[0]
                sheet.cell(row=rowCounter, column=8).value = None
                if not dropdownStyle:
                    sheet.cell(row=rowCounter, column=6).value = None
            elif not hasStyles and not hasSizes and hasColors:
                sheet.cell(row=rowCounter, column=8).value = selections[0]
                sheet.cell(row=rowCounter, column=7).value = None
                if not dropdownStyle:
                    sheet.cell(row=rowCounter, column=6).value = None
            print(selections)

            title = title.replace('\n', '')

            sheet.cell(row=rowCounter, column=1).value = title
            sheet.cell(row=rowCounter, column=2).value = brand
            sheet.cell(row=rowCounter, column=3).value = styleCount
            sheet.cell(row=rowCounter, column=4).value = sizeCount
            sheet.cell(row=rowCounter, column=5).value = colorCount

        timestamp = ('{:%m-%d-%Y %H:%M:%S}'.format(datetime.datetime.now()))
        sheet.cell(row=rowCounter, column=9).value = timestamp
        somethingsWrong = False
        for x in range(1,9):
            if sheet.cell(row=rowCounter, column=x).value != workbook["Actual"].cell(row=rowCounter, column=x).value:
                sheet.cell(row=rowCounter, column=x).fill = PatternFill(start_color='FFEE1111', end_color='FFEE1111', fill_type='solid')
                somethingsWrong = True
            else:
                sheet.cell(row=rowCounter, column=x).fill = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type='solid')
        if somethingsWrong:
            asin = url.replace("https://amazon.com/dp/", "")
            tempArray = [asin, selections]
            incorrectVariations.append(tempArray)
        workbook.save("Log.xlsx")
        rowCounter+=1
    print(incorrectVariations)
root = tk.Tk()

l1 = tk.Label(root, text="Start Row:", font=40)
l1.grid(row=0,column=0)

rowEntry1 = tk.Entry(root, font=40)
rowEntry1.grid(row=0,column=1)

l2 = tk.Label(root, text="End Row:", font=40)
l2.grid(row=1,column=0)

rowEntry2 = tk.Entry(root, font=40)
rowEntry2.grid(row=1,column=1)

button = tk.Button(root, text="Check Variations", font=40, command=lambda: scrapeAmazon(int(rowEntry1.get()),int(rowEntry2.get())))
button.grid(row=4,column=1)

root.mainloop()

